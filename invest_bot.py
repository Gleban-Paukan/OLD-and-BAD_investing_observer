import httplib2
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import os
import pandas as pd
import random
import sys
import telebot
import time as tm
from datetime import datetime, time, timedelta
from dateutil.parser import parse
from googleapiclient.discovery import build
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.styles import Font, Fill, PatternFill, NamedStyle, Side, Border
from requests.exceptions import ConnectionError, ReadTimeout
from telebot import types

import py_plot

bot = telebot.TeleBot('')
bot.parse_mode = 'html'
users_token = {}
users_bal = {}
users_login = {}
user = {}
ref_id = {}
ref_got = {}
time_id = {}
data = {}
moders = []
admins = []


@bot.message_handler(commands=['start'])
def welcome(message):
    try:
        if message.chat.type == 'private':
            with open('users_id.txt', 'r') as users_id:
                user_set = set()
                for i in users_id:
                    user_set.add(i.replace('\n', ''))
            if str(message.chat.id) not in user_set:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                markup.add(types.KeyboardButton('⁃ Разобраться в крипте '))
                markup.add(types.KeyboardButton('⁃ Пассивный доход '))
                markup.add(types.KeyboardButton('⁃ Заработок на реферальной программе '))
                markup.add(types.KeyboardButton('⁃ Ищу сообщество'))
                bot.send_message(message.chat.id,
                                 '''<b>
Привет! Добро пожаловать в сообщество криптоинвесторов!
Познакомимся? Всего три вопроса.
Что вас интересует, что для вас важно?
                                 </b>''', reply_markup=markup)
                bot.register_next_step_handler(message, expirience)
                with open('users_id.txt', 'a') as users_id:
                    users_id.write(str(message.chat.id) + '\n')
                with open('users_id.txt', 'r') as users_id:
                    user_set = set()
                    for i in users_id:
                        user_set.add(i.replace('\n', ''))
                with open('users_bal.txt', 'a') as non_set_bal:
                    non_set_bal.write(str(message.chat.id) + ' ' + '0' + '\n')
                with open('ref_got.txt', 'a') as non_set_ref:
                    non_set_ref.write(str(message.chat.id) + ' ' + '0' + '\n')
                with open('time_id.txt', 'a') as non_set_sub:
                    non_set_sub.write(str(message.chat.id) + ' ' + str((datetime.now()).strftime('%Y-%m-%d')) + '\n')
                if " " in message.text:
                    referrer_candidate = message.text.split()[1]
                    if str(message.chat.id) != referrer_candidate:
                        referer = referrer_candidate
                        non_set_ref = open('ref_id.txt', 'a')
                        non_set_ref.write(str(message.chat.id) + ' ' + referer + '\n')
                        non_set_ref.close()
                        ref_id[str(message.chat.id)] = referer

            else:
                if (bot.get_chat_member(-1001611319386, message.chat.id).status) == 'left':
                    markup = types.InlineKeyboardMarkup(row_width=1)
                    item5 = types.InlineKeyboardButton("🔗 Подписаться", url='https://t.me/LINK_IS_PRIVATE')
                    markup.add(item5)
                    bot.send_message(message.chat.id,
                                     '<b>Советуем зайти в наш чат. Там всегда дружелюбная атмосфера.</b>',
                                     parse_mode='html',
                                     reply_markup=markup)
                    tm.sleep(5)
                    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                    item0 = types.KeyboardButton("🚀 Регистрация")
                    item2 = types.KeyboardButton("♻️ Помощь")
                    item3 = types.KeyboardButton('⚙️ Профиль')
                    item7 = types.KeyboardButton('🤝 Реферальная программа')
                    item8 = types.KeyboardButton('💵 Инвестировать')
                    item9 = types.KeyboardButton('✅ Почему мы')
                    item123 = types.KeyboardButton('🎙 Общий чат')
                    markup.add(item0, item3)
                    markup.add(item2, item7)
                    markup.add(item8, item9)
                    markup.add(item123)
                    bot.send_message(message.chat.id,
                                     '<b>Добро пожаловать в мир инвестиций. Я являюсь посредником между командой профессиональных инвесторов и вами.\nПодробную инструкцию можно посмотреть, нажав на кнопку <i>«♻ <ins>️Помощь</ins> »</i>.\nПеред началом работы советуем указать ваш <i>USDT</i> токен во вкладке <i>«🚀 <ins>Регистрация </ins>»</i>. Его можно будет изменить.</b>'.format(
                                         message.from_user, bot.get_me()), parse_mode='html', reply_markup=markup)

                else:
                    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                    item0 = types.KeyboardButton("🚀 Регистрация")
                    item2 = types.KeyboardButton("♻️ Помощь")
                    item3 = types.KeyboardButton('⚙️ Профиль')
                    item7 = types.KeyboardButton('🤝 Реферальная программа')
                    item8 = types.KeyboardButton('💵 Инвестировать')
                    item9 = types.KeyboardButton('✅ Почему мы')
                    item123 = types.KeyboardButton('🎙 Общий чат')
                    markup.add(item0, item3)
                    markup.add(item2, item7)
                    markup.add(item8, item9)
                    markup.add(item123)
                    bot.send_message(message.chat.id,
                                     '<b>Добро пожаловать в мир инвестиций. Я являюсь посредником между командой профессиональных инвесторов и вами.\nПодробную инструкцию можно посмотреть, нажав на кнопку <i>«♻ <ins>️Помощь</ins> »</i>.\nПеред началом работы советуем указать ваш <i>USDT</i> токен во вкладке <i>«🚀 <ins>Регистрация </ins>»</i>. Его можно будет изменить.</b>'.format(
                                         message.from_user, bot.get_me()), parse_mode='html', reply_markup=markup)



    except Exception as er:
        print(er)


@bot.message_handler(content_types=['left_chat_member'])
def leave_memb(message):
    bot.delete_message(message.chat.id, message.id)


@bot.message_handler(content_types=['new_chat_members'])
def qweqwewq(message):
    bot.delete_message(message.chat.id, message.id)
    msg = bot.send_message(message.chat.id,
                           f'<ins><i>{message.from_user.first_name}</i></ins>, приветствуем вас в нашем сообществе!',
                           parse_mode='html')
    tm.sleep(60)
    bot.delete_message(msg.chat.id, msg.id)


@bot.message_handler(content_types=['text'])
def lalala(message):
    try:
        if message.chat.type == 'private':
            global user
            with open('verif_id.txt', 'r') as users_ver:
                user_ver = set()
                for i in users_ver:
                    user_ver.add(i.replace('\n', ''))
            with open('verif_id_2.txt', 'r') as users_ver_2:
                user_ver_2 = set()
                for i in users_ver_2:
                    user_ver_2.add(i.replace('\n', ''))
            with open('ref_got.txt') as file:
                for line in file:
                    key, *value = line.split()
                    ref_got[key] = value[0]
            with open('ref_id.txt') as file:
                for line in file:
                    key, *value = line.split()
                    ref_id[key] = value[0]
            with open('time_id.txt') as file:
                for line in file:
                    key, *value = line.split()
                    time_id[key] = value[0]
            with open('users_tok.txt') as file:
                for line in file:
                    key, *value = line.split()
                    users_token[key] = value[0]
            with open('users_bal.txt') as file:
                for line in file:
                    key, *value = line.split()
                    users_bal[key] = value[0]
            with open('users_log.txt') as file:
                for line in file:
                    key, *value = line.split()
                    users_login[key] = value[0]
            if message.text == '⚙️ Профиль':
                if str(message.chat.id) in users_login:
                    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                    item2 = types.KeyboardButton("➰ Изменить USDT-кошелёк")
                    item7 = types.KeyboardButton('🔙 Назад')
                    markup.add(item2)
                    markup.add(item7)
                    bot.send_message(message.chat.id, '<b>{0.first_name}, ваши данные: '.format(message.from_user,
                                                                                                bot.get_me()) + f'\n\n💳 Ваш USDT кошелёк: <code>{users_token[str(message.chat.id)]}</code>\nℹ️ Ваш ID телеграмм: <code>{message.chat.id}</code>\n\n💰 Ваш баланс: <ins><i>{users_bal[str(message.chat.id)]} USDT</i></ins>\n🥷🏿 Ваш логин: <ins><i>{users_login[str(message.chat.id)]}</i></ins>\n🕰 Ваш профиль зарегистрирован: <ins><i>{time_id[str(message.chat.id)]}</i></ins>\n💸 Партнёров приведено: <ins><i>{len(get_key_many(ref_id, str(message.chat.id)))} чел.</i></ins></b>',
                                     parse_mode='html', reply_markup=markup)
                else:
                    bot.send_message(message.chat.id,
                                     '<b>Пожалуйста, пройдите регистрацию. Нажмите на <i>«🚀 <ins>Регистрация</ins>»</i></b>',
                                     parse_mode='html')
            elif message.text == '🚀 Регистрация':
                if str(message.chat.id) not in users_token:
                    bot.send_message(message.chat.id,
                                     '<b>Пожалуйста, введите ваш USDT TRC-20.\n\n <a href="https://telegra.ph/Kak-uznat-svoj-koshelyok-USDT-TRC20-06-03-3">Как узнать кошелёк USDT-TRC20</a></b>',
                                     parse_mode='html')
                    bot.register_next_step_handler(message, token_remember)
                elif str(message.chat.id) not in users_login:
                    bot.send_message(message.chat.id,
                                     '<b>Пожалуйста, введите ваш логин ( псевдоним ) одним словом. Его нельзя будет изменить.</b>',
                                     parse_mode='html')
                    bot.register_next_step_handler(message, login_remember)
                else:
                    bot.send_message(message.chat.id,
                                     '<b>Вы уже зарегистрированы. Для просмотра ваших данных, нажмите <i>«👨🏻 <ins>Профиль</ins>»</i></b>',
                                     parse_mode='html')
            elif message.text == '➖ Вывести деньги':
                bot.send_message(message.chat.id, '<b>Введите, сколько хотите вывести.</b>', parse_mode='html')
                bot.register_next_step_handler(message, cash_out)
            elif message.text == 'В главное меню':
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                item0 = types.KeyboardButton("🚀 Регистрация")
                item2 = types.KeyboardButton("♻️ Помощь")
                item3 = types.KeyboardButton('⚙️ Профиль')
                item7 = types.KeyboardButton('🤝 Реферальная программа')
                item8 = types.KeyboardButton('💵 Инвестировать')
                item9 = types.KeyboardButton('✅ Почему мы')
                item123 = types.KeyboardButton('🎙 Общий чат')
                item10 = types.KeyboardButton('/admin')
                markup.add(item3)
                markup.add(item2, item7)
                markup.add(item8, item9)
                markup.add(item0)
                markup.add(item123)
                markup.add(item10)
                bot.send_message(message.chat.id,
                                 '<b>Хозяин барин.</b>'.format(
                                     message.from_user, bot.get_me()), parse_mode='html', reply_markup=markup)
            elif message.text == '🎙 Общий чат':
                bot.send_message(message.chat.id,
                                 '<b>Переходи в наш <i> <a href="https://t.me/LINK_IS_PRIVATE">чат</a></i>!</b>',
                                 parse_mode='html')
            elif message.text == '/admin' and (str(message.chat.id) in admins):
                user[message.chat.id] = ''
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                item0 = types.KeyboardButton("🚀 Регистрация")
                item2 = types.KeyboardButton("♻️ Помощь")
                item3 = types.KeyboardButton('⚙️ Профиль')
                item7 = types.KeyboardButton('🤝 Реферальная программа')
                item8 = types.KeyboardButton('💵 Инвестировать')
                item9 = types.KeyboardButton('✅ Почему мы')
                item123 = types.KeyboardButton('🎙 Общий чат')
                markup.add(item0, item3)
                markup.add(item2, item7)
                markup.add(item8, item9)
                markup.add(item123)
                item14 = types.KeyboardButton('Памятка')
                item10 = types.KeyboardButton('Найти человека по ID телеграмм')
                item11 = types.KeyboardButton('Найти человека по логину')
                item12 = types.KeyboardButton('Изменить баланс ( по ID телеграмма )')
                item13 = types.KeyboardButton('Изменить баланс ( по логину )')
                markup.add(item14)
                markup.add(item10)
                markup.add(item11)
                markup.add(item12)
                markup.add(item13)
                markup.add('Количество пользователей в боте')
                markup.add('/cash_out')
                bot.send_message(message.chat.id, '<b>Узнаю.</b>', parse_mode='html', reply_markup=markup)
            elif message.text == 'Количество пользователей в боте':
                with open('ready_id.txt', 'r') as users_id:
                    user_red = set()
                    for i in users_id:
                        user_red.add(i.replace('\n', ''))
                bot.send_message(message.chat.id, len(users_bal))
                count = 0
                for i in users_bal:
                    if users_bal[i] != '0':
                        count += 1
                bot.send_message(message.chat.id, f'Из них с балансом:<code> {count}</code>')
                bot.send_message(message.chat.id, f'Из них прошло воронку: <code> {len(user_red)}</code>')
            elif message.text == '/cash_out':
                if str(message.chat.id) in admins:
                    bot.send_message(message.chat.id, '<b>Введите, по какому проценту считать.</b>', parse_mode='html')
                    bot.register_next_step_handler(message, pereschet)
            elif message.text == 'Памятка':
                bot.send_message(message.chat.id,
                                 '<b>• <ins><i>/admin</i></ins> Показывает админ панель с функциями: <i>простомотр информации о пользователе, а также изменение баланса пользователя</i>\n\n• <ins><i>/cash_out</i></ins> спрашивает, месячный процент и высчитывает весь доход пользователя за месяц. Отправляет всё в чат "Вывод процентов" и составляет excel-таблицу\n\n• <ins><i>/send</i></ins> спрашивает ID пользователя и через пробел сообщение. Сообщение можно делать на абзацы, главное с пробелом.\n\n• <ins><i>/send_all</i></ins> в одной строке саму команду и через пробел сообщение.\nПример: <i>/send_all Вывод денег произойдёт в 12:00 по МСК!</i></b>',
                                 parse_mode='html')
            elif message.text == 'инвест':
                markup = types.InlineKeyboardMarkup(row_width=1)
                item10 = types.InlineKeyboardButton("Переходи в наш чат!", url='https://t.me/LINK_IS_PRIVATE')
                markup.add(item10)
                with open('photo_2022-06-05_16-22-43.jpg', 'rb') as file:
                    bot.send_photo(message.chat.id, file, caption='''<b>
Почему мы?

🔰 NAME IS PRIVATE – это управляющая инвестиционная компания, которая позволяет получать дополнительный доход безопасно и легко.
Мы работаем с июля 2021 года и показываем успешные результаты.

В ряде наших преимуществ:

☑️ Надёжность и прозрачность.
Наша компания фиксирует все договорённости и обеспечивает безопасность своим инвесторам.
Мы предоставляем видеоотчёты с процесса торгов, тем самым гарантируя прозрачность. Формат содержит несколько повесток: отчётность, новости, вопросы и ответ.

☑️ Опыт.
В нашей команде работают трейдеры и аналитики, обладающие опытом более 8 лет, которые знают как приносить результаты. Наши трейдеры сотрудничают с топ 10 трейдерами по торговле в странах СНГ на бирже Binance. Они постоянно 24/7 изучают рынок и подбирают самые оптимальные монеты для инвестирования.

Доверие – ключевой смысл бизнеса, так же доверие ключевой фактор семьи. Доверьтесь нам и станьте частью нашей семьи!
О своей семье мы заботимся в первую очередь.</b>'
    ''', parse_mode='html', reply_markup=markup)
                bot.send_message(message.chat.id,
                                 '<b>✅ <ins><i>Антикризисное предложение</i></ins>\n\n💸 Мы представляем вашему вниманию <ins>путь к пассивному заработку</ins>. Мы — <i>Криптоинвесторы <ins>LINK_IS_PRIVATE </ins></i>\n🎉 Переходи по ссылке и узнай, как официально получить ставку 40% годовых!</b>',
                                 parse_mode='html')
            elif message.text == '✅ Почему мы':
                bot.send_message(message.chat.id,
                                 '''<b>
Почему мы?

🔰 NAME IS PRIVATE – это управляющая инвестиционная компания, которая позволяет получать дополнительный доход безопасно и легко.
Мы работаем с июля 2021 года и показываем успешные результаты.
Торгуем на мировой бирже Binance, которая является лидером на рынке.
Люди доверяют нам свои инвестиции, а мы показываем людям, что можно выйти на пассивный заработок честно и без обмана.

В ряде наших преимуществ:

☑️ <ins>Надёжность и прозрачность.</ins>
Наша компания фиксирует все договорённости и обеспечивает безопасность своим инвесторам.
Мы предоставляем видеоотчёты с процесса торгов, тем самым гарантируя прозрачность. Формат содержит несколько повесток: отчётность, новости, вопросы и ответ.

☑️ <ins>Опыт.</ins>
В нашей команде работают трейдеры и аналитики, обладающие опытом более 8 лет, которые знают как приносить результаты. Наши трейдеры сотрудничают с топ 10 трейдерами по торговле в странах СНГ на бирже Binance. Они постоянно 24/7 изучают рынок и подбирают самые оптимальные монеты для инвестирования.

☑️ <ins>Бонусы.</ins>
У нашей компании есть разовое вознаграждение (реферальный бонус) за нового привлечённого инвестора в виде 3% от суммы внесенного депозита и 1% с прибыли каждый месяц. 

Доверие – ключевой смысл бизнеса, так же доверие ключевой фактор семьи. Доверьтесь нам и станьте частью нашей семьи!
О своей семье мы заботимся в первую очередь.</b>
                                 ''', parse_mode='html')
            elif message.text == '📉 Понизить баланс у пользователя':
                bot.send_message(message.chat.id, '<b>На сколько рублей?</b>', parse_mode='html')
                bot.register_next_step_handler(message, minus_balance_by_login)
            elif message.text == '📈 Повысить баланс у пользователя':
                bot.send_message(message.chat.id, '<b>На сколько рублей?</b>', parse_mode='html')
                bot.register_next_step_handler(message, add_balance_by_login)
            elif '/send_all' in message.text and str(message.chat.id) in admins:
                users_id = set()
                non_set = open('users_id.txt', 'r')
                for i in non_set:
                    users_id.add(i.strip())
                non_set.close()
                for user in users_id:
                    try:
                        bot.send_message(user, message.text[message.text.find(' '):])
                    except Exception as er:
                        print(er)
            elif '/send' in str(message.text) and str(message.chat.id) in admins:
                bot.send_message(message.chat.id, 'Введите ID и сообщение через пробел')
                bot.register_next_step_handler(message, send)
            elif message.text == '🔝 Вывести реферальные деньги':
                if ref_got[str(message.chat.id)] != 0:
                    ref_out(message)
                else:
                    bot.send_message(message.chat.id, '<b>Пока что на вашем счете ноль. Приведите кого-то.</b>',
                                     parse_mode='html')
            elif message.text == 'Изменить баланс ( по логину )':
                bot.send_message(message.chat.id, '<b>Введите логин человека</b>', parse_mode='html')
                bot.register_next_step_handler(message, change_balanse_by_login)
            elif message.text == '💔 Понизить баланс у пользователя':
                bot.send_message(message.chat.id, '<b>На сколько рублей?</b>', parse_mode='html')
                bot.register_next_step_handler(message, minus_balance_by_id)
            elif message.text == '💛 Повысить баланс у пользователя':
                bot.send_message(message.chat.id, '<b>На сколько рублей?</b>', parse_mode='html')
                bot.register_next_step_handler(message, add_balance_by_id)
            elif message.text == 'Изменить баланс ( по ID телеграмма )':
                bot.send_message(message.chat.id, '<b>Введите ID человека</b>', parse_mode='html')
                bot.register_next_step_handler(message, change_balanse_by_id)
            elif message.text == 'Найти человека по ID телеграмм':
                bot.send_message(message.chat.id, '<b>Введите ID пользователя.</b>', parse_mode='html')
                bot.register_next_step_handler(message, find_user_by_ID)
            elif message.text == 'Найти человека по логину':
                bot.send_message(message.chat.id, '<b>Введите логин пользователя.</b>', parse_mode='html')
                bot.register_next_step_handler(message, find_user_by_login)
            elif message.text == '➕ Пополнить счёт':
                bot.send_message(message.chat.id,
                                 '<b>Для пополнения счёта перевердите монеты на кошелёк <code>5379331848:AAGA05o5sIL8SK623wt2aRWFAw7RlRGUfgs</code> и напишите <i><a href="https://t.me/LINK_IS_PRIVATE">модератору бота</a></i>. Будьте готовы прислать <ins>txid-код</ins>.</b>',
                                 parse_mode='html')
            elif message.text == '➰ Изменить USDT-кошелёк':
                bot.send_message(message.chat.id, '<b>Введите ваш USDT-токен.</b>', parse_mode='html')
                bot.register_next_step_handler(message, token_change)
            elif message.text == '🤝 Реферальная программа':
                bot.send_message(message.chat.id,
                                 f'<b>По условиям нашей реферальной программы вы можете получить по 3% разового бонуса от инвестиций вашего реферала и 1% будете получать ежемесячно от суммы приведенных инвестиций. Количество рефералов неограничено.\n\nВаша реферальная ссылка: t.me/LINK_IS_PRIVATE?start={message.chat.id}</b>',
                                 parse_mode='html')
            elif message.text == '♻️ Помощь':
                bot.send_message(message.chat.id, 'https://telegra.ph/Vopros-otvet-06-13')
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                item1 = types.KeyboardButton('❌  Закрыть обращение')
                item2 = types.KeyboardButton('🔙 Главное меню')  # 🔙
                markup.add(item1, item2)
                bot.send_message(message.chat.id,
                                 '📩 Напишите ваше обращение ниже. Мы ответим в течении суток.\n\n❌️ Для закрытия обращения, нажмите кнопку «<b>Закрыть обращение</b>»',
                                 reply_markup=markup)
                bot.register_next_step_handler(message, peresil)
            elif message.text == '🧮 Калькулятор прибыли':
                bot.send_message(message.chat.id,
                                 '<b>Введите сумму, относительно которой хотите рассчитать прибыль</b>',
                                 parse_mode='html')
                bot.register_next_step_handler(message, calculate_money)
            elif message.text == '❓ Как инвестировать':
                bot.send_message(message.chat.id,
                                 '<b>Прочитайте <i><a href="https://telegra.ph/Kak-investirovat-06-03">cтатью</a></i> </b>',
                                 parse_mode='html')
            elif message.text == '📈 Начать инвестировать':
                # if str(message.chat.id) in user_ver:
                bot.send_message(message.chat.id,
                                 f'<b>Напишите нашему <i><a href="https://t.me/{random.choice(moders)}">модератору</a></i>.</b>',
                                 disable_web_page_preview=True, parse_mode='html')
                # else:
                #     msg = bot.send_message(message.chat.id, '<b>Вам необходимо подтвердить соглашение.</b>', parse_mode='html')
                #     markup = types.InlineKeyboardMarkup(row_width=1)
                #     item6 = types.InlineKeyboardButton('✅ Согласен', callback_data='ver_user')
                #     markup.add(item6)
                #     with open('договор_доверительного_управления_с_данными_ооо.doc', 'rb') as file:
                #         bot.send_document(message.chat.id, file, reply_markup=markup)
                #     time.sleep(30)
                #     bot.delete_message(msg.chat.id, msg.id)
            elif message.text == '💵 Инвестировать':
                if str(message.chat.id) in user_ver_2:
                    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                    item2 = types.KeyboardButton("🔙 Назад")
                    item3 = types.KeyboardButton('❓ Как инвестировать')
                    item4 = types.KeyboardButton('📈 Начать инвестировать')
                    markup.add(item3, item4)
                    markup.add(item2)
                    bot.send_message(message.chat.id,
                                     '''<b>
                                 ✅Вы можете инвестировать, чтобы зарабатывать пассивно 3-5% в месяц.

💰Минимальный депозит - 100 USDT.
💰Максимальный депозит - без ограничений.</b>''', parse_mode='html',
                                     reply_markup=markup)
                else:
                    msg = bot.send_message(message.chat.id, '<b>Вам необходимо подтвердить соглашение.</b>',
                                           parse_mode='html')
                    markup = types.InlineKeyboardMarkup(row_width=1)
                    item6 = types.InlineKeyboardButton('✅ Согласен', callback_data='ver_user_2')
                    markup.add(item6)
                    with open('Пользовательское_соглашение_на_телеграмм_бот.doc', 'rb') as file:
                        bot.send_document(message.chat.id, file, reply_markup=markup)
                    tm.sleep(30)
                    bot.delete_message(msg.chat.id, msg.id)
            elif str(message.text).lower() == 'топ':
                array = []
                for i in users_bal:
                    array.append(users_bal[i])
                for i in range(len(array) - 1, 0, -1):
                    if array[i] == '0':
                        array.remove(array[i])
                counter = 1
                text = ''
                i = 0
                array = sorted(array, reverse=True)
                while i < len(array):
                    if len(get_key_many(users_bal, array[i])) != 1:
                        for j in range(len(get_key_many(users_bal, array[i]))):
                            text += f'{counter + j} — {users_login[get_key_many(users_bal, array[i])[j]]}, {users_bal[get_key_many(users_bal, array[i])[j]]}\n'
                        i += len(get_key_many(users_bal, array[i]))
                        counter += len(get_key_many(users_bal, array[i]))
                    else:
                        text += f'{counter} — {users_login[get_key_many(users_bal, array[i])[0]]}, {users_bal[get_key_many(users_bal, array[i])[0]]}\n'
                        i += 1
                    counter += 1
                bot.send_message(message.chat.id, f'<b>{text}</b>', parse_mode='html')
            elif message.text == 'маркетинг':
                bot.send_message(message.chat.id,
                                 '<b>Введи, какие теги вам нужны. Необхоимо вводить каждый тег с новой строки. Пример:\n\ntg_01\ntg__991\nGenWake</b>',
                                 parse_mode='html')
                bot.register_next_step_handler(message, marketing)
            elif message.text == '🔙 Назад':
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                item0 = types.KeyboardButton("🚀 Регистрация")
                item2 = types.KeyboardButton("♻️ Помощь")
                item3 = types.KeyboardButton('⚙️ Профиль')
                item7 = types.KeyboardButton('🤝 Реферальная программа')
                item8 = types.KeyboardButton('💵 Инвестировать')
                item9 = types.KeyboardButton('✅ Почему мы')
                item123 = types.KeyboardButton('🎙 Общий чат')
                markup.add(item0, item3)
                markup.add(item2, item7)
                markup.add(item8, item9)
                markup.add(item123)
                bot.send_message(message.chat.id,
                                 '<b>Возвращено в главное меню</b>'.format(
                                     message.from_user, bot.get_me()), parse_mode='html', reply_markup=markup)
        else:
            if message.chat.id == -1001654956289:
                if '/send' in str(message.text):
                    id = str(message.text)[
                         len('/send') + 1:len('/send') + 1 + (str(message.text)[len('/send') + 1:].find(' '))]
                    bot.send_message(int(id), str(message.text)[
                                              len('/send') + 2 + (str(message.text)[len('/send') + 1:].find(' ')):])
                else:
                    if message.reply_to_message.forward_from != None:
                        bot.send_message(message.reply_to_message.forward_from.id, message.text)
                    else:
                        bot.send_message(message.chat.id, 'У пользователя анонимка')
    except Exception as er:
        print(er)


@bot.callback_query_handler(func=lambda call: True)
def callback_inline(call):
    try:
        if call.message:
            if call.data == 'check_sub':
                if (bot.get_chat_member(-1001611319386, call.message.chat.id).status) == 'left':
                    markup = types.InlineKeyboardMarkup(row_width=1)
                    item5 = types.InlineKeyboardButton("🔗 Подписаться", url='https://t.me/LINK_IS_PRIVATE')
                    item6 = types.InlineKeyboardButton('✔️ Подписался', callback_data='check_sub')
                    markup.add(item5, item6)
                    bot.send_message(call.message.chat.id,
                                     '<b>Прверка не пройдена. Подпишитесь на канал, пройдя по кнопке ниже.</b>',
                                     parse_mode='html',
                                     reply_markup=markup)
                else:
                    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                    item0 = types.KeyboardButton("🚀 Регистрация")
                    item2 = types.KeyboardButton("♻️ Помощь")
                    item3 = types.KeyboardButton('⚙️ Профиль')
                    item7 = types.KeyboardButton('🤝 Реферальная программа')
                    item8 = types.KeyboardButton('💵 Инвестировать')
                    item9 = types.KeyboardButton('✅ Почему мы')
                    item123 = types.KeyboardButton('🎙 Общий чат')
                    markup.add(item0, item3)
                    markup.add(item2, item7)
                    markup.add(item8, item9)
                    markup.add(item123)
                    bot.send_message(call.message.chat.id,
                                     '<b>Добро пожаловать в мир инвестиций. Я являюсь посредником между командой профессиональных инвесторов и вами.\nПодробную инструкцию можно посмотреть, нажав на кнопку <i>«📜<ins> Помощь</ins>»</i>.\nПеред началом работы советуем ваш USDT токен во вкладке <i>«👨🏻 <ins>Профиль</ins>»</i>. Его можно будет изменить.</b>'.format(
                                         call.message.from_user, bot.get_me()), parse_mode='html', reply_markup=markup)
            elif call.data == 'ver_user':
                bot.send_message(call.message.chat.id,
                                 f'<b>Напишите нашему <i><a href="https://t.me/{random.choice(moders)}">модератору</a></i>.</b>',
                                 disable_web_page_preview=True, parse_mode='html')
                with open('verif_id.txt', 'a') as users_id:
                    users_id.write(str(call.message.chat.id) + '\n')
            elif call.data == 'ver_user_2':
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                item2 = types.KeyboardButton("🔙 Назад")
                item3 = types.KeyboardButton('❓ Как инвестировать')
                item4 = types.KeyboardButton('📈 Начать инвестировать')
                markup.add(item3, item4)
                markup.add(item2)
                bot.send_message(call.message.chat.id,
                                 '''<b>
✅Вы можете инвестировать, чтобы зарабатывать пассивно 3-5% в месяц.

💰Минимальный депозит - 100 USDT.
💰Максимальный депозит - без ограничений.</b>''', parse_mode='html',
                                 reply_markup=markup)
                with open('verif_id_2.txt', 'a') as users_id:
                    users_id.write(str(call.message.chat.id) + '\n')
            bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
    except Exception as er:
        print(er)


@bot.message_handler(content_types=['contact'])
def lalala(message):
    try:
        google_sheets_registration([data[str(message.chat.id)]])
        print(message)
        bot.send_message(message.chat.id, '<b>Отлично! Мы свяжемся с вами в течение дня!</b>')
        if (bot.get_chat_member(-1001611319386, message.chat.id).status) == 'left':
            markup = types.InlineKeyboardMarkup(row_width=1)
            item5 = types.InlineKeyboardButton("🔗 Подписаться", url='https://t.me/LINK_IS_PRIVATE')
            markup.add(item5)
            bot.send_message(message.chat.id, '<b>Советуем зайти в наш чат. Там всегда дружелюбная атмосфера.</b>',
                             parse_mode='html',
                             reply_markup=markup)
            tm.sleep(5)
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item0 = types.KeyboardButton("🚀 Регистрация")
            item2 = types.KeyboardButton("♻️ Помощь")
            item3 = types.KeyboardButton('⚙️ Профиль')
            item7 = types.KeyboardButton('🤝 Реферальная программа')
            item8 = types.KeyboardButton('💵 Инвестировать')
            item9 = types.KeyboardButton('✅ Почему мы')
            item123 = types.KeyboardButton('🎙 Общий чат')
            markup.add(item0, item3)
            markup.add(item2, item7)
            markup.add(item8, item9)
            markup.add(item123)
            bot.send_message(message.chat.id,
                             '<b>Добро пожаловать в мир инвестиций. Я являюсь посредником между командой профессиональных инвесторов и вами.\nПодробную инструкцию можно посмотреть, нажав на кнопку <i>«♻ <ins>️Помощь</ins> »</i>.\nПеред началом работы советуем указать ваш <i>USDT</i> токен во вкладке <i>«🚀 <ins>Регистрация </ins>»</i>. Его можно будет изменить.</b>'.format(
                                 message.from_user, bot.get_me()), parse_mode='html', reply_markup=markup)

        else:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item0 = types.KeyboardButton("🚀 Регистрация")
            item2 = types.KeyboardButton("♻️ Помощь")
            item3 = types.KeyboardButton('⚙️ Профиль')
            item7 = types.KeyboardButton('🤝 Реферальная программа')
            item8 = types.KeyboardButton('💵 Инвестировать')
            item9 = types.KeyboardButton('✅ Почему мы')
            item123 = types.KeyboardButton('🎙 Общий чат')
            markup.add(item0, item3)
            markup.add(item2, item7)
            markup.add(item8, item9)
            markup.add(item123)
            bot.send_message(message.chat.id,
                             '<b>Добро пожаловать в мир инвестиций. Я являюсь посредником между командой профессиональных инвесторов и вами.\nПодробную инструкцию можно посмотреть, нажав на кнопку <i>«♻ <ins>️Помощь</ins> »</i>.\nПеред началом работы советуем указать ваш <i>USDT</i> токен во вкладке <i>«🚀 <ins>Регистрация </ins>»</i>. Его можно будет изменить.</b>'.format(
                                 message.from_user, bot.get_me()), parse_mode='html', reply_markup=markup)
    except Exception as er:
        print(er)


def expirience(message):
    data[str(message.chat.id)] = [message.text]
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('⁃ Пробовал торговать'))
    markup.add(types.KeyboardButton('⁃ Новичок'))
    bot.send_message(message.chat.id,
                     '''<b>
                     Какой у вас опыт торговли криптой?
                     </b>''', reply_markup=markup)
    bot.register_next_step_handler(message, investing)


def investing(message):
    data[str(message.chat.id)].append(message.text)
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('⁃ Никогда не пробовал'))
    markup.add(types.KeyboardButton('⁃ Пробовал по минимуму'))
    markup.add(types.KeyboardButton('⁃ Собрал небольшой портфель'))
    markup.add(types.KeyboardButton('⁃ Торгую постоянно'))
    markup.add(types.KeyboardButton('⁃ Другое'))
    bot.send_message(message.chat.id,
                     '''<b>
Есть ли уже опыт получения заработка на другого рода инвестициях, в акции и ценные бумаги, например?
                     </b>''', reply_markup=markup)
    bot.register_next_step_handler(message, final)


def final(message):
    data[str(message.chat.id)].append(message.text)
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('⁃ Да', request_contact=True))
    markup.add(types.KeyboardButton('⁃ Пока нет'))
    bot.send_message(message.chat.id,
                     '''<b>
Супер, спасибо! Мы команда трейдеров и аналитиков, которая торгует на криптовалюте и успешно генерирует прибыль. Готовы вас проконсультировать. Что скажете, пообщаемся?
                     </b>''', reply_markup=markup)
    bot.register_next_step_handler(message, final_FORSURE)


def final_FORSURE(message):
    with open('ready_id.txt', 'a') as users_id:
        users_id.write(str(message.chat.id) + '\n')
    # print(message.contact.phone_number)
    if type(message.chat.username) != str:
        user_name = 'Пользователь его не указал'
    else:
        user_name = message.chat.username
    data[str(message.chat.id)].append(user_name)
    data[str(message.chat.id)].append(message.chat.id)
    if message.content_type == 'contact':
        data[str(message.chat.id)].append(message.contact.phone_number)
        data[str(message.chat.id)].append(message.contact.first_name)
        data[str(message.chat.id)].append(message.contact.last_name)
        bot.send_message(message.chat.id, '<b>Отлично! Мы свяжимся с вами в течение дня.</b>')
        bot.send_message(-730859322,
                         f'''
Зашёл новый пользователь. 
Его ответы на вопросы:
Важность: <i>{data[str(message.chat.id)][0]}</i>;
Опыт: <i>{data[str(message.chat.id)][1]}</i>;
Опыт в заработке: <i>{data[str(message.chat.id)][2]}</i>;
Username: @<i>{data[str(message.chat.id)][3]}</i>;
ID аккаунта: <code>{data[str(message.chat.id)][4]}</code>
Номер телефона: <code>{data[str(message.chat.id)][5]}</code>;
Имя на аккаунте: <i>{data[str(message.chat.id)][6]}</i>.
''')
    else:
        bot.send_message(-730859322,
                         f'''
Зашёл новый пользователь. 
Его ответы на вопросы:
Важность: <i>{data[str(message.chat.id)][0]}</i>;
Опыт: <i>{data[str(message.chat.id)][1]}</i>;
Опыт в заработке: <i>{data[str(message.chat.id)][2]}</i>;
Username: @<i>{data[str(message.chat.id)][3]}</i>
ID аккаунта: <code>{data[str(message.chat.id)][4]}</code>
        ''')
    google_sheets_registration([data[str(message.chat.id)]])
    if (bot.get_chat_member(-1001611319386, message.chat.id).status) == 'left':
        markup = types.InlineKeyboardMarkup(row_width=1)
        item5 = types.InlineKeyboardButton("🔗 Подписаться", url='https://t.me/LINK_IS_PRIVATE')
        markup.add(item5)
        bot.send_message(message.chat.id, '<b>Советуем зайти в наш чат. Там всегда дружелюбная атмосфера.</b>',
                         parse_mode='html',
                         reply_markup=markup)
        tm.sleep(5)
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item0 = types.KeyboardButton("🚀 Регистрация")
        item2 = types.KeyboardButton("♻️ Помощь")
        item3 = types.KeyboardButton('⚙️ Профиль')
        item7 = types.KeyboardButton('🤝 Реферальная программа')
        item8 = types.KeyboardButton('💵 Инвестировать')
        item9 = types.KeyboardButton('✅ Почему мы')
        item123 = types.KeyboardButton('🎙 Общий чат')
        markup.add(item0, item3)
        markup.add(item2, item7)
        markup.add(item8, item9)
        markup.add(item123)
        bot.send_message(message.chat.id,
                         '<b>Добро пожаловать в мир инвестиций. Я являюсь посредником между командой профессиональных инвесторов и вами.\nПодробную инструкцию можно посмотреть, нажав на кнопку <i>«♻ <ins>️Помощь</ins> »</i>.\nПеред началом работы советуем указать ваш <i>USDT</i> токен во вкладке <i>«🚀 <ins>Регистрация </ins>»</i>. Его можно будет изменить.</b>'.format(
                             message.from_user, bot.get_me()), parse_mode='html', reply_markup=markup)

    else:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item0 = types.KeyboardButton("🚀 Регистрация")
        item2 = types.KeyboardButton("♻️ Помощь")
        item3 = types.KeyboardButton('⚙️ Профиль')
        item7 = types.KeyboardButton('🤝 Реферальная программа')
        item8 = types.KeyboardButton('💵 Инвестировать')
        item9 = types.KeyboardButton('✅ Почему мы')
        item123 = types.KeyboardButton('🎙 Общий чат')
        markup.add(item0, item3)
        markup.add(item2, item7)
        markup.add(item8, item9)
        markup.add(item123)
        bot.send_message(message.chat.id,
                         '<b>Добро пожаловать в мир инвестиций. Я являюсь посредником между командой профессиональных инвесторов и вами.\nПодробную инструкцию можно посмотреть, нажав на кнопку <i>«♻ <ins>️Помощь</ins> »</i>.\nПеред началом работы советуем указать ваш <i>USDT</i> токен во вкладке <i>«🚀 <ins>Регистрация </ins>»</i>. Его можно будет изменить.</b>'.format(
                             message.from_user, bot.get_me()), parse_mode='html', reply_markup=markup)


def peresil(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item0 = types.KeyboardButton("🚀 Регистрация")
    item2 = types.KeyboardButton("♻️ Помощь")
    item3 = types.KeyboardButton('⚙️ Профиль')
    item7 = types.KeyboardButton('🤝 Реферальная программа')
    item8 = types.KeyboardButton('💵 Инвестировать')
    item9 = types.KeyboardButton('✅ Почему мы')
    item123 = types.KeyboardButton('🎙 Общий чат')
    markup.add(item0, item3)
    markup.add(item2, item7)
    markup.add(item8, item9)
    markup.add(item123)
    if message.text == '🔙 Главное меню':
        bot.send_message(message.chat.id, '❇️ Вы перешли в <b>Главное меню</b>', reply_markup=markup)
    elif message.text != '❌  Закрыть обращение':
        bot.forward_message(-1001654956289, message.chat.id, message.message_id)
        if type(message.chat.username) != str:
            user_name = 'Пользователь его не указал'
        else:
            user_name = message.chat.username
        bot.send_message(-1001654956289, f'ID пользователя <code>{message.chat.id}</code>, username: @{user_name}')
        bot.send_message(message.chat.id, 'Ваше обращение открыто. Ожидайте ответа модератора.')
        bot.register_next_step_handler(message, peresil_1)
    else:
        bot.send_message(message.chat.id, 'Ваше обращение закрыто.', reply_markup=markup)


def peresil_1(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item0 = types.KeyboardButton("🚀 Регистрация")
    item2 = types.KeyboardButton("♻️ Помощь")
    item3 = types.KeyboardButton('⚙️ Профиль')
    item7 = types.KeyboardButton('🤝 Реферальная программа')
    item8 = types.KeyboardButton('💵 Инвестировать')
    item9 = types.KeyboardButton('✅ Почему мы')
    item123 = types.KeyboardButton('🎙 Общий чат')
    markup.add(item0, item3)
    markup.add(item2, item7)
    markup.add(item8, item9)
    markup.add(item123)
    if message.text == '🔙 Главное меню':
        bot.send_message(message.chat.id, '❇️ Вы перешли в <b>Главное меню</b>', reply_markup=markup)
    elif message.text != '❌  Закрыть обращение':
        bot.forward_message(-1001654956289, message.chat.id, message.message_id)
        bot.register_next_step_handler(message, peresil_1)
    else:
        bot.send_message(message.chat.id, 'Ваше обращение закрыто.', reply_markup=markup)
        bot.send_message(-1001654956289,
                         f'Обращение пользователя <ins><i>{message.from_user.first_name}</i></ins> закрыто.')


def marketing(message):
    tags = str(message.text).split('\n')
    py_plot.plot(tags)
    with open('marketing.png', 'rb') as file:
        bot.send_photo(message.chat.id, file)


def send(message):
    try:
        if ' ' in message.text:
            bot.send_message(str(message.text)[0:str(message.text).index(' ') + 1],
                             str(message.text)[str(message.text).index(' ') + 1:])
            bot.send_message(message.chat.id, 'Сообщение отправил.')
        else:
            bot.send_message(message.chat.id, 'Так и бота можно сломать, нужен пробел.')
    except:
        pass


def get_key(x, value):
    for k, v in x.items():
        if v == value:
            return k


def get_key_many(x, value):
    q = []
    for k, v in x.items():
        if v == value:
            q.append(k)
    return q


def find_user_by_ID(message):
    if message.text in users_token:
        bot.send_message(message.chat.id,
                         '<b>Данные человека: ' + f'\n\n💳 USDT кошелёк: <code>{users_token[str(message.text)]}</code>\nℹ️ ID телеграмм: <code>{message.text}</code>\n\n💰 Баланс: <ins><i>{users_bal[str(message.text)]} USD</i></ins>\n🥷🏿 Логин: <ins><i>{users_login[str(message.text)]}</i></ins>\n🕰 Профиль зарегистрирован: <ins><i>{time_id[str(message.text)]}\n</i></ins>💸 Партнёрская программа принесла: <ins><i>{ref_got[str(message.text)]} USD</i></ins></b>',
                         parse_mode='html')
    else:
        bot.send_message(message.chat.id, '<b>Не смог такого найти :(</b>', parse_mode='html')


def ref_out(message):
    if float(ref_got[str(message.chat.id)]) != 0:
        with open('ref_got.txt') as file:
            for line in file:
                key, *value = line.split()
                ref_got[key] = value[0]
        bot.send_message(-713833994,
                         f'<b>Поступил запрос на вывод <i><ins>РЕФЕРАЛЬНЫХ</ins></i> средств.\nЗапросил: <ins>{str(message.chat.id)} </ins> ( логин <ins>{users_login[str(message.chat.id)]}</ins>).\n\nСумма: <code>{round(float(ref_got[str(message.chat.id)]), 3)}</code>.\nКошелёк: <code> {users_token[str(message.chat.id)]} </code></b>',
                         parse_mode='html')
        with open('ref_got.txt', 'r') as f:
            old_data = f.read()
        new_data = old_data.replace(str(message.chat.id) + ' ' + ref_got[str(message.chat.id)],
                                    str(message.chat.id) + ' ' + '0')
        with open('ref_got.txt', 'w') as f:
            f.write(new_data)
    else:
        bot.send_message(message.chat.id, '<b>На вашем реферальном счету 0 рублей.</b>', parse_mode='html')


def pereschet(message):
    with open('ref_id.txt') as file:
        for line in file:
            key, *value = line.split()
            ref_id[key] = value[0]
    with open('ref_got.txt') as file:
        for line in file:
            key, *value = line.split()
            ref_got[key] = value[0]
    with open('time_id.txt') as file:
        for line in file:
            key, *value = line.split()
            time_id[key] = value[0]
    with open('users_tok.txt') as file:
        for line in file:
            key, *value = line.split()
            users_token[key] = value[0]
    with open('users_bal.txt') as file:
        for line in file:
            key, *value = line.split()
            users_bal[key] = value[0]
    with open('users_log.txt') as file:
        for line in file:
            key, *value = line.split()
            users_login[key] = value[0]
    normal_users = {}
    for i in users_bal:
        if users_bal[i] != '0':
            normal_users[i] = users_bal[i]
    values = ref_id.values()
    procent = float(message.text) * 0.01
    couter = 1
    logins = []
    time = []
    summ_us = []
    token = []
    summ_ref_one = []
    summ_ref_all = []
    summ_all = []
    for i in normal_users:
        if i not in values:
            if (abs(parse(time_id[i]) - datetime.now())) < timedelta(days=30):
                bot.send_message(-690723715,
                                 f'<b><i><ins>Запрос номер {str(couter)}</ins></i>\n\nПользователь {i} (логин {users_login[i]})\n\nЗарегистрирован: {time_id[i]}\n\nСумма: <code> {round(float(users_bal[i]) * abs(float(str((parse(time_id[i]) - datetime.now()).days))) * (procent / 30), 3)}</code>\n\nКошелёк: <code> {users_token[i]} </code> </b>',
                                 parse_mode='html')
                logins.append(users_login[i])
                time.append(time_id[i])
                summ_us.append(round(
                    float(users_bal[i]) * abs(float(str((parse(time_id[i]) - datetime.now()).days))) * (procent / 30),
                    3))
                token.append(users_token[i])
                summ_ref_one.append(0)
                summ_ref_all.append(0)
                summ_all.append(round(
                    float(users_bal[i]) * abs(float(str((parse(time_id[i]) - datetime.now()).days))) * (procent / 30),
                    3))
            else:
                bot.send_message(-690723715,
                                 f'<b><i><ins>Запрос номер {str(couter)}</ins></i>\n\nПользователь {i} (логин {users_login[i]})\n\nЗарегистрирован: {time_id[i]}\n\nСумма: <code> {round(float(users_bal[i]) * procent, 3)}</code>\n\nКошелёк: <code> {users_token[i]} </code> </b>',
                                 parse_mode='html')
                logins.append(users_login[i])
                time.append(time_id[i])
                summ_us.append(round(float(users_bal[i]) * procent, 3))
                token.append(users_token[i])
                summ_ref_one.append(0)
                summ_ref_all.append(0)
                summ_all.append(round(float(users_bal[i]) * procent, 3))
        else:
            all_summ = 0
            refers = get_key_many(ref_id, i)
            summ = 0
            for j in range(len(refers)):
                if (abs(parse(time_id[refers[j]]) - datetime.now())) > timedelta(days=30):
                    summ = round(float(users_bal[refers[j]]) * 0.01, 3)
                if summ != 0:
                    all_summ += summ
            if (abs(parse(time_id[i]) - datetime.now())) < timedelta(days=30):
                bot.send_message(-690723715,
                                 f'<b><i><ins>Запрос номер {str(couter)}</ins></i>\n\nПользователь {i} (логин {users_login[i]})\n\nЗарегистрирован: {time_id[i]}\n\nСумма: <code> {round(float(users_bal[i]) * abs(float(str((parse(time_id[i]) - datetime.now()).days))) * (procent / 30), 3)}</code>\nСумма за рефералов: <code> {all_summ}</code>\nСумма за рефералов (единоразово): <code> {ref_got[i]}</code>\nСумма по всем рефералам: <code>{float(ref_got[i]) + all_summ}</code>\n-------------------------\nИтого: <code> {round(float(users_bal[i]) * abs(float(str((parse(time_id[i]) - datetime.now()).days))) * (procent / 30), 3) + float(ref_got[i]) + all_summ} </code>\n\nКошелёк: <code> {users_token[i]} </code> </b>',
                                 parse_mode='html')
                logins.append(users_login[i])
                time.append(time_id[i])
                summ_us.append(round(
                    float(users_bal[i]) * abs(float(str((parse(time_id[i]) - datetime.now()).days))) * (procent / 30),
                    3))
                token.append(users_token[i])
                summ_ref_one.append(all_summ)
                summ_ref_all.append(float(ref_got[i]) + all_summ)
                summ_all.append(round(
                    float(users_bal[i]) * abs(float(str((parse(time_id[i]) - datetime.now()).days))) * (procent / 30),
                    3) + float(ref_got[i]) + all_summ)
            else:
                bot.send_message(-690723715,
                                 f'<b><i><ins>Запрос номер {str(couter)}</ins></i>\n\nПользователь {i} (логин {users_login[i]})\n\nЗарегистрирован: {time_id[i]}\n\nСумма: <code> {round(float(users_bal[i]) * procent, 3)}\nСумма за рефералов: <code> {all_summ}</code>\nСумма за рефералов (единоразово): <code> {ref_got[i]}</code>\nСумма по всем рефералам: <code>{float(ref_got[i]) + all_summ}</code>\n-------------------------\nИтого: <code> {round(float(users_bal[i]) * procent, 3) + float(ref_got[i]) + all_summ} </code>\n\nКошелёк: <code> {users_token[i]} </code> </b>',
                                 parse_mode='html')

                logins.append(users_login[i])
                time.append(time_id[i])
                summ_us.append(round(float(users_bal[i]) * procent, 3))
                token.append(users_token[i])
                summ_ref_one.append(all_summ)
                summ_ref_all.append(float(ref_got[i]) + all_summ)
                summ_all.append(round(float(users_bal[i]) * procent, 3) + float(ref_got[i]) + all_summ)
        couter += 1
        df = pd.DataFrame(
            {
                'Логины': logins,
                'Время регистрации': time,
                'Сумма': summ_us,
                'Кошелёк': token,
                'Сумма по рефералам': summ_ref_one,
                'Сумма по рефералам (Одноразово)': summ_ref_all,
                'Сумма полная': summ_all}
        )
        # Save Styler Object for Later
        df.to_excel('./final.xlsx')
        book = openpyxl.load_workbook('final.xlsx')
        sheet = book.active
        sheet.column_dimensions['C'].width = len('Время регистрации')
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 35
        sheet.column_dimensions['F'].width = 22
        sheet.column_dimensions['G'].width = 25
        sheet.column_dimensions['H'].width = 20
        for q in range(len(logins)):
            sheet[f'B{str(q + 2)}'].style = 'Good'
        for q in range(len(logins)):
            sheet[f'C{str(q + 2)}'].style = 'Bad'
        for q in range(len(logins)):
            sheet[f'D{str(q + 2)}'].style = 'Linked Cell'
        for q in range(len(logins)):
            sheet[f'F{str(q + 2)}'].style = 'Input'
        for q in range(len(logins)):
            sheet[f'D{str(q + 2)}'].style = 'Neutral'
        font = Font(b=True, color='00DD00')
        fill2 = PatternFill('solid', fgColor='FFD700')
        for q in range(len(logins)):
            sheet[f'G{str(q + 2)}'].font = font
            sheet[f'G{str(q + 2)}'].fill = fill2

        for q in range(len(logins)):
            sheet[f'H{str(q + 2)}'].style = 'Note'
        book.save('./final.xlsx')
        for i in ref_got:
            with open('ref_got.txt', 'r') as f:
                old_data = f.read()
            new_data = old_data.replace(str(i) + ' ' + ref_got[i],
                                        i + ' ' + '0')
            with open('ref_got.txt', 'w') as f:
                f.write(new_data)
    with open('final.xlsx', 'rb') as file:
        bot.send_document(-690723715, file)


def color(row):
    if row.isnull().values.any() == True:
        return ['background-color: red'] * len(row)
    return [''] * len(row)


def find_user_by_login(message):
    values = users_login.values()
    if message.text in values:
        ID = get_key(users_login, message.text)
        bot.send_message(message.chat.id,
                         '<b>Данные человека: ' + f'\n\n💳 USDT кошелёк: <code>{users_token[ID]}</code>\nℹ️ ID телеграмм: <code>{ID}</code>\n\n💰 Баланс: <ins><i>{users_bal[ID]} RUB</i></ins>\n🥷🏿 Логин: <ins><i>{users_login[ID]}</i></ins>\n🕰 Профиль зарегистрирован: <ins><i>{time_id[ID]}\n</i></ins>💸 Партнёрская программа принесла: <ins><i>{ref_got[ID]} USD</i></ins></b>',
                         parse_mode='html')
    else:
        bot.send_message(message.chat.id, '<b>Такого логина не нашёл :(</b>', parse_mode='html')


def token_change(message):
    with open('users_tok.txt') as file:
        for line in file:
            key, *value = line.split()
            users_token[key] = value[0]
    if ' ' in message.text or len(message.text) > 100 or not match(str(message.text).lower()):
        bot.send_message(message.chat.id, '<b>Вы указали неправильный кошелёк. Пожалуйста, попробуйте ещё раз.</b>',
                         parse_mode='html')
    else:
        with open('users_tok.txt', 'r') as f:
            old_data = f.read()
        new_data = old_data.replace(str(message.chat.id) + ' ' + users_token[str(message.chat.id)],
                                    str(message.chat.id) + ' ' + message.text)
        with open('users_tok.txt', 'w') as f:
            f.write(new_data)
        bot.send_message(message.chat.id, '<b>Ваш токен успешно изменён!</b>', parse_mode='html')


def token_remember(message):
    if ' ' in message.text or len(message.text) > 100 or not match(str(message.text).lower()):
        bot.send_message(message.chat.id,
                         '<b>Вы указали неправильный кошелёк. Пожалуйста, попробуйте ещё раз (Нажмите на <i>«🚀 <ins>Регистрация</ins>»</i>).</b>',
                         parse_mode='html')
    else:
        with open('users_tok.txt', 'a') as users_tok:
            users_tok.write(str(message.chat.id) + ' ' + message.text + '\n')
        if str(message.chat.id) not in users_login:
            bot.send_message(message.chat.id,
                             '<b>Также, пожалуйста, введите ваш логин ( псевдоним ) одним словом. Его нельзя будет изменить.</b>',
                             parse_mode='html')
            bot.register_next_step_handler(message, login_remember)


def login_remember(message):
    with open('users_log.txt') as file:
        for line in file:
            key, *value = line.split()
            users_login[key] = value[0]
    values = users_login.values()
    if ' ' in message.text or len(message.text) > 100 or not match(str(message.text).lower()):
        bot.send_message(message.chat.id,
                         '<b>Вы указали неправильный Логин. Пожалуйста, попробуйте ещё раз (Нажмите на <i>«🚀 <ins>Регистрация</ins>»</i>).</b>',
                         parse_mode='html')
    elif message.text in values:
        bot.send_message(message.chat.id, '<b>Такой логин уже используется. Выберите другой</b>', parse_mode='html')
    else:
        with open('users_log.txt', 'a') as users_tok:
            users_tok.write(str(message.chat.id) + ' ' + message.text + '\n')
        bot.send_message(message.chat.id, '<b>Ваш логин добавлен!</b>', parse_mode='html')
        with open('users_log.txt') as file:
            for line in file:
                key, *value = line.split()
                users_login[key] = value[0]
        with open('users_tok.txt') as file:
            for line in file:
                key, *value = line.split()
                users_token[key] = value[0]
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item2 = types.KeyboardButton("➰ Изменить USDT-кошелёк")
        item7 = types.KeyboardButton('🔙 Назад')
        markup.add(item2)
        markup.add(item7)
        bot.send_message(message.chat.id, '<b>{0.first_name}, ваши данные: '.format(message.from_user,
                                                                                    bot.get_me()) + f'\n\n💳 Ваш USDT кошелёк: <code>{users_token[str(message.chat.id)]}</code>\nℹ️ Ваш ID телеграмм: <code>{message.chat.id}</code>\n\n💰 Ваш баланс: <ins><i>{users_bal[str(message.chat.id)]} USDT</i></ins>\n🥷🏿 Ваш логин: <ins><i>{users_login[str(message.chat.id)]}</i></ins>\n🕰 Ваш профиль зарегистрирован: <ins><i>{time_id[str(message.chat.id)]}</i></ins>\n💸 Партнёрская программа принесла: <ins><i>{str(round(float(ref_got[str(message.chat.id)]), 2))} USDT</i></ins></b>',
                         parse_mode='html', reply_markup=markup)
        bot.send_message(-730859322,
                         f'<b>Зарегистрирован новый пользователь. Его токоен: <code> {users_token[str(message.chat.id)]}.</code>\nЕго логин: <ins><i>{users_login[str(message.chat.id)]}</i></ins></b>',
                         parse_mode='html')


def minus_balance_by_id(message):
    if str(message.chat.id) in admins:
        with open('users_bal.txt', 'r') as f:
            old_data = f.read()
        new_data = old_data.replace(user[message.chat.id] + ' ' + users_bal[user[message.chat.id]],
                                    user[message.chat.id] + ' ' + str(
                                        float(users_bal[user[message.chat.id]]) - float(message.text)))
        with open('users_bal.txt', 'w') as f:
            f.write(new_data)
        with open('users_bal.txt') as file:
            for line in file:
                key, *value = line.split()
                users_bal[key] = value[0]
        bot.send_message(message.chat.id,
                         f'<b>Баланс изменён. Ныншений баланс поль зователя {user[message.chat.id]} : <ins> {users_bal[user[message.chat.id]]} </ins></b>',
                         parse_mode='html')


def minus_balance_by_login(message):
    if str(message.chat.id) in admins:
        with open('users_bal.txt', 'r') as f:
            old_data = f.read()
        new_data = old_data.replace(user[message.chat.id] + ' ' + users_bal[user[message.chat.id]],
                                    user[message.chat.id] + ' ' + str(
                                        float(users_bal[user[message.chat.id]]) - float(message.text)))
        with open('users_bal.txt', 'w') as f:
            f.write(new_data)
        with open('users_bal.txt') as file:
            for line in file:
                key, *value = line.split()
                users_bal[key] = value[0]
        bot.send_message(message.chat.id,
                         f'<b>Баланс изменён. Ныншений баланс пользователя {user[message.chat.id]} : <ins> {users_bal[user[message.chat.id]]} </ins></b>',
                         parse_mode='html')


def add_balance_by_id(message):
    if str(message.chat.id) in admins:
        with open('ref_id.txt') as file:
            for line in file:
                key, *value = line.split()
                ref_id[key] = value[0]
        with open('users_bal.txt', 'r') as f:
            old_data = f.read()
        new_data = old_data.replace(user[message.chat.id] + ' ' + users_bal[user[message.chat.id]],
                                    user[message.chat.id] + ' ' + str(
                                        float(message.text) + float(users_bal[user[message.chat.id]])))
        with open('users_bal.txt', 'w') as f:
            f.write(new_data)
        with open('users_bal.txt') as file:
            for line in file:
                key, *value = line.split()
                users_bal[key] = value[0]
        if str(user[message.chat.id]) not in ref_id:
            bot.send_message(message.chat.id,
                             f'<b>Баланс изменён. Ныншений баланс пользователя {user[message.chat.id]} : <ins> {users_bal[user[message.chat.id]]} </ins></b>',
                             parse_mode='html')
        else:
            with open('ref_got.txt', 'r') as f:
                old_data = f.read()
            new_data = old_data.replace(ref_id[user[message.chat.id]] + ' ' + ref_got[ref_id[user[message.chat.id]]],
                                        ref_id[user[message.chat.id]] + ' ' + str(
                                            float(float(message.text) * 0.03) + float(
                                                ref_got[ref_id[user[message.chat.id]]])))
            with open('ref_got.txt', 'w') as f:
                f.write(new_data)
            with open('users_bal.txt') as file:
                for line in file:
                    key, *value = line.split()
                    users_bal[key] = value[0]
            bot.send_message(message.chat.id,
                             f'<b>Баланс изменён. Ныншений баланс пользователя {user[message.chat.id]} : <ins> {users_bal[user[message.chat.id]]} </ins>\nА ещё этот пользователь был рефером <ins><i>{users_login[ref_id[user[message.chat.id]]]}</i></ins>.</b>',
                             parse_mode='html')


def add_balance_by_login(message):
    if str(message.chat.id) in admins:
        with open('ref_id.txt') as file:
            for line in file:
                key, *value = line.split()
                ref_id[key] = value[0]
        with open('ref_got.txt') as file:
            for line in file:
                key, *value = line.split()
                ref_got[key] = value[0]
        with open('users_bal.txt', 'r') as f:
            old_data = f.read()
        new_data = old_data.replace(user[message.chat.id] + ' ' + users_bal[user[message.chat.id]],
                                    user[message.chat.id] + ' ' + str(
                                        float(message.text) + float(users_bal[user[message.chat.id]])))
        with open('users_bal.txt', 'w') as f:
            f.write(new_data)
        with open('users_bal.txt') as file:
            for line in file:
                key, *value = line.split()
                users_bal[key] = value[0]
        if str(user[message.chat.id]) not in ref_id:
            bot.send_message(message.chat.id,
                             f'<b>Баланс изменён. Ныншений баланс пользователя {user[message.chat.id]} : <ins> {users_bal[user[message.chat.id]]} </ins></b>',
                             parse_mode='html')
        else:
            with open('ref_got.txt', 'r') as f:
                old_data = f.read()
            new_data = old_data.replace(ref_id[user[message.chat.id]] + ' ' + ref_got[ref_id[user[message.chat.id]]],
                                        ref_id[user[message.chat.id]] + ' ' + str(
                                            float(float(message.text) * 0.03) + float(
                                                ref_got[ref_id[user[message.chat.id]]])))
            with open('ref_got.txt', 'w') as f:
                f.write(new_data)
            with open('users_bal.txt') as file:
                for line in file:
                    key, *value = line.split()
                    users_bal[key] = value[0]
            bot.send_message(message.chat.id,
                             f'<b>Баланс изменён. Ныншений баланс пользователя {user[message.chat.id]} : <ins> {users_bal[user[message.chat.id]]} </ins>\nА ещё этот пользователь был рефером <ins><i>{users_login[ref_id[user[message.chat.id]]]}</i></ins>.</b>',
                             parse_mode='html')


def change_balanse_by_id(message):
    if message.text in users_token:
        user[message.chat.id] = message.text
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item2 = types.KeyboardButton("💔 Понизить баланс у пользователя")
        item3 = types.KeyboardButton("💛 Повысить баланс у пользователя")
        item4 = types.KeyboardButton("В главное меню")
        markup.add(item3)
        markup.add(item2)
        markup.add(item4)
        bot.send_message(message.chat.id, f'<b>Нашёл такого. Его нынешний баланс: {users_bal[message.text]}. </b>',
                         parse_mode='html', reply_markup=markup)
    else:
        bot.send_message(message.chat.id, '<b>Не нашёл такого.</b>', parse_mode='html')


def change_balanse_by_login(message):
    user[message.chat.id] = get_key(users_login, message.text)
    if user[message.chat.id] in users_token:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item2 = types.KeyboardButton("📉 Понизить баланс у пользователя")
        item3 = types.KeyboardButton("📈 Повысить баланс у пользователя")
        item4 = types.KeyboardButton("В главное меню")
        markup.add(item3)
        markup.add(item2)
        markup.add(item4)
        bot.send_message(message.chat.id,
                         f'<b>Нашёл такого. Его нынешний баланс: {users_bal[user[message.chat.id]]}. </b>',
                         parse_mode='html', reply_markup=markup)
    else:
        bot.send_message(message.chat.id, '<b>Не нашёл такого.</b>', parse_mode='html')


def cash_out(message):
    try:
        if float(message.text) > float(users_bal[str(message.chat.id)]):
            bot.send_message(message.chat.id, '<b>К сожалению, у вас столько нет.</b>', parse_mode='html')
        else:
            bot.send_message(-713833994,
                             f'<i> Пользователь {message.chat.id} (логин <b><ins>{users_login[str(message.chat.id)]}</ins></b>) запросил вывод на <code>{message.text}</code>.\n\nЕго кошелёк: <code>{users_token[str(message.chat.id)]}</code></i>',
                             parse_mode='html')
    except Exception as er:
        bot.send_message(message.chat.id, '<b>Пожалуйста, ввидете число.</b>', parse_mode='html')
        print(er)


def calculate_money(message):
    try:
        num = float(message.text)
        bot.send_message(message.chat.id,
                         f'<b>В среднем наша прибыль в месяц составляет 4%. \n\n<i>Прибыль за месяц: <ins>{str(num * 0.04)} RUB</ins>\nПрибыль за год: <ins>{str(num * 0.04 * 12)} RUB</ins></i></b>',
                         parse_mode='html')

    except ValueError:
        bot.send_message(message.chat.id, '<b>Вы неправильно указали сумму инвестиций.</b>', parse_mode='html')


def match(text):
    alphabet = set('абвгдеёжзийклмнопрстуфхцчшщъыьэюя')
    return alphabet.isdisjoint(text.lower())


def get_service_sacc():
    scopes = ['https://www.googleapis.com/auth/spreadsheets']

    creds_service = ServiceAccountCredentials.from_json_keyfile_name('creds.json', scopes).authorize(httplib2.Http())
    return build('sheets', 'v4', http=creds_service)


def google_sheets_registration(values):
    sheet = get_service_sacc().spreadsheets()

    sheet_id = "1M_yvvYre05nMgC_5I4a7DIMQGXj-AHiPt38FVdXa27o"

    sheet.values().append(
        spreadsheetId=sheet_id,
        range="Воронка продаж",
        valueInputOption="RAW",
        body={'values': values}).execute()


try:
    bot.infinity_polling(timeout=10, long_polling_timeout=5)
except (ConnectionError, ReadTimeout) as e:
    sys.stdout.flush()
    os.execv(sys.argv[0], sys.argv)
else:
    bot.infinity_polling(timeout=10, long_polling_timeout=5)
